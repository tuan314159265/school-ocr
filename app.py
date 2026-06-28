"""
Giao diện Streamlit — School Registration OCR to Excel.

Chức năng:
    A. Mở file Excel:           load file .xlsx có sẵn để tiếp tục.
    B. Upload & Scan:           chụp OCR từ ảnh, tự động thêm vào danh sách.
    C. Data Editor:             xem/sửa/xoá dòng dữ liệu.
    D. Lưu file:                tải xuống file Excel tổng hợp.
"""

from __future__ import annotations

import json
import sys as _sys
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

from ocr_engine import OCREngine, SCHEMA_KEYS

# ---------------------------------------------------------------------------
# Cấu hình
# ---------------------------------------------------------------------------

# Lấy API key — ưu tiên Streamlit Secrets (cloud), fallback config.json (local)
def _get_api_key_and_model() -> tuple[str, str]:
    """Lấy API key và model từ Streamlit Secrets hoặc config.json."""
    # Thử từ Streamlit Secrets trước
    try:
        api_key = st.secrets.get("api_key", "")
        model = st.secrets.get("model", "gemini-3.1-flash-lite")
        if api_key:
            return api_key, model
    except Exception:
        pass

    # Fallback: config.json cạnh app.py / .exe
    if getattr(_sys, "frozen", False):
        config_dir = Path(_sys.argv[0]).parent
    else:
        config_dir = Path(__file__).parent

    config_file = config_dir / "config.json"
    if config_file.exists():
        try:
            cfg = json.loads(config_file.read_text(encoding="utf-8"))
            return cfg.get("api_key", ""), cfg.get("model", "gemini-3.1-flash-lite")
        except Exception:
            pass

    return "", "gemini-3.1-flash-lite"

# Mapping key -> tên cột Excel hiển thị theo thứ tự (xuất ra)
# key "" là cột trống
EXCEL_COLUMNS: list[tuple[str, str]] = [
    ("_stt", "STT"),
    ("ho_ten_hoc_sinh", "Họ và tên học sinh"),
    ("ngay_sinh", "Ngày tháng năm sinh"),
    ("gioi_tinh", "Giới tính"),
    ("noi_sinh", "Nơi sinh"),
    ("dan_toc", "Dân tộc"),
    ("noi_cu_tru", "Địa chỉ"),
    ("", ""),
    ("ho_ten_cha", "Họ tên cha"),
    ("ho_ten_me", "Họ tên mẹ"),
]

# Ánh xạ ngược — tên cột Excel -> key (khi import)
HEADER_TO_KEY: dict[str, str] = {
    "Họ và tên học sinh": "ho_ten_hoc_sinh",
    "Ngày tháng năm sinh": "ngay_sinh",
    "Giới tính": "gioi_tinh",
    "Dân tộc": "dan_toc",
    "Nơi sinh": "noi_sinh",
    "Địa chỉ": "noi_cu_tru",
    "Họ tên cha": "ho_ten_cha",
    "Họ tên mẹ": "ho_ten_me",
    "Họ và tên": "ho_ten_hoc_sinh",
    "Ngày sinh": "ngay_sinh",
    "Nơi thường trú": "noi_cu_tru",
    "Nơi cư trú": "noi_cu_tru",
    "Cha": "ho_ten_cha",
    "Mẹ": "ho_ten_me",
    "Họ tên người cha": "ho_ten_cha",
    "Họ tên người mẹ": "ho_ten_me",
}

# Cột hiển thị trong data_editor
EDITOR_COLUMNS = SCHEMA_KEYS  # 10 keys

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def confidence_emoji(value: str) -> str:
    """Chuyển _confidence thành emoji."""
    mapping = {"ok": "✅", "low": "⚠️", "error": "❌"}
    return mapping.get(value, "❓")


def is_duplicate(records: list[dict], new_record: dict) -> bool:
    """
    Kiểm tra new_record có trùng với record nào trong danh sách không.
    So sánh theo 2 cấp: họ tên+ngày sinh -> họ tên+cha+mẹ.
    """
    for rec in records:
        # Cấp 1: họ tên + ngày sinh
        new_name = new_record.get("ho_ten_hoc_sinh", "").strip().lower()
        old_name = rec.get("ho_ten_hoc_sinh", "").strip().lower()
        new_dob = new_record.get("ngay_sinh", "").strip()
        old_dob = rec.get("ngay_sinh", "").strip()
        if new_name and old_name and new_dob and old_dob:
            if new_name == old_name and new_dob == old_dob:
                return True

        # Cấp 2: họ tên + họ tên cha + họ tên mẹ
        new_father = new_record.get("ho_ten_cha", "").strip().lower()
        old_father = rec.get("ho_ten_cha", "").strip().lower()
        new_mother = new_record.get("ho_ten_me", "").strip().lower()
        old_mother = rec.get("ho_ten_me", "").strip().lower()
        if new_name and old_name and new_father and old_father and new_mother and old_mother:
            if new_name == old_name and new_father == old_father and new_mother == old_mother:
                return True

    return False


def read_excel_to_records(file) -> list[dict]:
    """Đọc file Excel, chuẩn hoá header, trả về danh sách records."""
    df = pd.read_excel(file)
    df = df.rename(columns=HEADER_TO_KEY)
    valid_cols = [c for c in df.columns if c in SCHEMA_KEYS]
    records = []
    for _, row in df.iterrows():
        record = {k: str(row.get(k, "")) for k in SCHEMA_KEYS}
        record["_confidence"] = "ok"
        record["_filename"] = "imported"
        records.append(record)
    return records


def records_to_dataframe(records: list[dict]) -> pd.DataFrame:
    """Chuyển records thành DataFrame với đủ cột."""
    df = pd.DataFrame(records)
    for col in SCHEMA_KEYS + ["_confidence", "_filename"]:
        if col not in df.columns:
            df[col] = ""
    return df


def create_excel(df: pd.DataFrame) -> BytesIO:
    """
    Tạo file Excel từ DataFrame theo thứ tự cột EXCEL_COLUMNS.
    Tự động thêm STT, thêm cột trống.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách học sinh"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Ghi header theo EXCEL_COLUMNS
    for col_idx, (key, header_name) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Ghi dữ liệu theo từng dòng
    for row_idx in range(len(df)):
        excel_row = row_idx + 2  # dòng 1 là header
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
            if key == "_stt":
                value = row_idx + 1  # STT tự động
            elif key == "":
                value = ""  # cột trống
            else:
                value = df.iloc[row_idx].get(key, "")

            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Auto-adjust độ rộng cột
    for col_idx, (key, header_name) in enumerate(EXCEL_COLUMNS, start=1):
        if key == "":
            ws.column_dimensions[get_column_letter(col_idx)].width = 3
            continue
        # Ước lượng độ rộng
        max_len = len(header_name)
        if key != "_stt" and key in df.columns:
            col_data_len = df[key].astype(str).map(len).max() if len(df) > 0 else 0
            max_len = max(max_len, col_data_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Khởi tạo — đọc config, kết nối Gemini
# ---------------------------------------------------------------------------

st.set_page_config(page_title="School OCR", layout="wide")
st.title("School Registration OCR — Nhập liệu tự động")

# Đọc API key từ Secrets (cloud) hoặc config.json (local)
API_KEY, MODEL_NAME = _get_api_key_and_model()

# Cache engine trong session_state để không gọi test_connection() mỗi lần re-render
if "engine" not in st.session_state or st.session_state.get("_last_api_key") != API_KEY:
    engine: OCREngine | None = None
    engine_ready = False

    if API_KEY:
        engine = OCREngine(API_KEY)
        engine.update_model(MODEL_NAME)
        engine_ready = engine.test_connection()

    st.session_state["engine"] = engine
    st.session_state["_engine_ready"] = engine_ready
    st.session_state["_last_api_key"] = API_KEY

engine = st.session_state.get("engine")
engine_ready = st.session_state.get("_engine_ready", False)

if not engine_ready:
    st.error(
        "Chưa có API key hoặc key không hợp lệ.\n\n"
        "- **Dùng local:** tạo file **config.json** cùng thư mục với:\n"
        '  `{"api_key": "AIzaSy...", "model": "gemini-3.1-flash-lite"}`\n\n'
        "- **Dùng Streamlit Cloud:** vào Settings → Secrets, thêm:\n"
        "  ```\n"
        "  api_key = \"AIzaSy...\"\n"
        "  model = \"gemini-2.0-flash\"\n"
        "  ```"
    )
    st.stop()

# Khởi tạo session state
if "records" not in st.session_state:
    st.session_state["records"] = []
if "open_filename" not in st.session_state:
    st.session_state["open_filename"] = "danh_sach_hoc_sinh.xlsx"

# ---------------------------------------------------------------------------
# A. Mở file Excel có sẵn
# ---------------------------------------------------------------------------

with st.expander("Mở file Excel có sẵn để tiếp tục", expanded=False):
    st.markdown(
        "Tải lên file **.xlsx** cũ — dữ liệu sẽ được nạp vào để xem, sửa hoặc bổ sung thêm."
    )

    open_file = st.file_uploader("Chọn file Excel (.xlsx)", type=["xlsx"], key="file_open")

    col_open1, col_open2 = st.columns([1, 3])
    with col_open1:
        if st.button("Mở file", type="primary", disabled=not open_file):
            try:
                opened_records = read_excel_to_records(open_file)
                st.session_state["records"] = opened_records
                st.session_state["open_filename"] = open_file.name
                st.success(f"Đã mở **{open_file.name}** — {len(opened_records)} học sinh.")
                time.sleep(1)
                st.rerun()
            except Exception as exc:
                st.error(f"Lỗi đọc file: {exc}")
    with col_open2:
        total = len(st.session_state["records"])
        st.caption(
            f"Hiện tại: **{total}** học sinh" if total > 0 else "Chưa có dữ liệu."
        )

# ---------------------------------------------------------------------------
# B. Upload & Scan
# ---------------------------------------------------------------------------

col_upload, col_status = st.columns([3, 1])

with col_upload:
    uploaded_files = st.file_uploader(
        "Chọn ảnh Giấy Khai Sinh / Trích lục Khai Sinh",
        type=["jpg", "jpeg", "png"],
        accept_multiple_files=True,
    )

with col_status:
    st.metric("Số học sinh", len(st.session_state["records"]))

if st.button("Bắt đầu quét dữ liệu", type="primary", disabled=not uploaded_files):
    engine.update_model(MODEL_NAME)

    progress_bar = st.progress(0, text="Đang chuẩn bị...")
    status_text = st.empty()

    total = len(uploaded_files)

    # Xử lý song song bằng ThreadPool — mỗi ảnh chạy độc lập
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Mở ảnh trước (thread-safe)
    images = []
    for f in uploaded_files:
        try:
            img = Image.open(f)
            images.append((img, f.name))
        except Exception:
            images.append((None, f.name))

    results: list[dict | None] = [None] * total

    with ThreadPoolExecutor(max_workers=min(total, 5)) as executor:
        future_map = {}
        for idx, (img, fname) in enumerate(images):
            if img is None:
                # Ảnh lỗi -> ghi nhận lỗi luôn
                result = {k: "" for k in SCHEMA_KEYS}
                result["_confidence"] = "error"
                result["_filename"] = fname
                results[idx] = result
                continue

            future = executor.submit(engine.process_image, img)
            future_map[future] = (idx, fname)

        skip_count = 0
        new_count = 0
        done_count = 0

        for future in as_completed(future_map):
            idx, fname = future_map[future]
            done_count += 1
            status_text.info(f"Đang xử lý: {fname} ({done_count}/{total})")

            try:
                result = future.result()
            except Exception:
                result = {k: "" for k in SCHEMA_KEYS}
                result["_confidence"] = "error"
                result["_filename"] = fname

            if is_duplicate(st.session_state["records"], result):
                skip_count += 1
            else:
                st.session_state["records"].append(result)
                new_count += 1

            progress_bar.progress(done_count / total, text=f"Đã xử lý {done_count}/{total}")

    status_text.success(f"Hoàn thành: thêm {new_count}, bỏ qua {skip_count} trùng.")
    time.sleep(1.5)
    st.rerun()

# ---------------------------------------------------------------------------
# C. Data Editor
# ---------------------------------------------------------------------------

st.divider()
st.subheader("Danh sách học sinh")

records = st.session_state.get("records", [])
if records:
    df = records_to_dataframe(records)

    df_display = df.copy()
    if "_confidence" in df_display.columns:
        df_display["_confidence"] = df_display["_confidence"].apply(confidence_emoji)

    display_columns = ["_confidence", "_filename"] + EDITOR_COLUMNS
    display_columns = [c for c in display_columns if c in df_display.columns]

    # Tạo mapping header cho data_editor từ EXCEL_COLUMNS
    editor_header_map = {}
    for key, header_name in EXCEL_COLUMNS:
        if key and key != "_stt":
            editor_header_map[key] = header_name

    edited = st.data_editor(
        df_display[display_columns],
        width="stretch",
        hide_index=True,
        column_config={
            "_confidence": st.column_config.TextColumn("Trạng thái", disabled=True, width="small"),
            "_filename": st.column_config.TextColumn("Tập tin", disabled=True, width="medium"),
            **{key: st.column_config.TextColumn(header_name, width="medium") for key, header_name in EXCEL_COLUMNS if key and key != "_stt"},
        },
        num_rows="dynamic",
    )

    if edited is not None and not edited.equals(df_display[display_columns]):
        updated_records = []
        for _, row in edited.iterrows():
            record = {k: str(row.get(k, "")) for k in SCHEMA_KEYS}
            record["_filename"] = str(row.get("_filename", ""))
            filename = row.get("_filename", "")
            matching = df[df["_filename"] == filename]
            record["_confidence"] = (
                matching.iloc[0]["_confidence"] if not matching.empty else "error"
            )
            updated_records.append(record)
        st.session_state["records"] = updated_records

    # Xoá dòng cuối
    col_del, col_exp = st.columns([1, 5])
    with col_del:
        if st.button("Xoá dòng cuối", type="secondary"):
            if st.session_state["records"]:
                st.session_state["records"].pop()
                st.rerun()

else:
    st.info("Chưa có dữ liệu. Hãy mở file Excel cũ hoặc upload ảnh để quét OCR.")

# ---------------------------------------------------------------------------
# D. Lưu file
# ---------------------------------------------------------------------------

st.divider()

if records:
    df_export = records_to_dataframe(records)

    col1, col2 = st.columns([1, 3])
    with col1:
        excel_buf = create_excel(df_export)
        base_name = st.session_state.get("open_filename", "danh_sach_hoc_sinh.xlsx")
        stem = Path(base_name).stem
        save_name = f"{stem}_updated.xlsx"

        st.download_button(
            label="Lưu file Excel",
            data=excel_buf,
            file_name=save_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    with col2:
        st.caption(f"Tổng số: {len(df_export)} học sinh")
else:
    st.info("Chưa có dữ liệu để lưu.")
